import requests
import statistics
import math
import csv
import os
import openpyxl
from openpyxl import Workbook
while True:
    index_lstt=['NIFTY50','NEXT50','AUTO','NIFTYBANK','FMCG','IT','FINSERV','MEDIA','METAL','PHARMA','PVTBK','PSUBK','REALTY','NIFTY200','NIFTY500','BASKET']
    print('\nINDICIES\n--------------------------------------------------------\n\n1-NIFTY50\n2-NEXT_50\n3-AUTO\n4-BANK\n5-FMCG\n6-IT\n7-FINANCIAL-SERVICES\n8-MEDIA\n9-METAL\n10-PHARMA\n11-PRIVATE-BANK\n12-PSU-BANK\n13-REALTY\n14-NIFTY200\n15-NIFTY500\n16-BASKET\n\n--------------------------------------------------------\n')
    s1=input('SELECT>>> ')
    s2=int(s1)-1
    wb1q= openpyxl.load_workbook(index_lstt[s2]+'.xlsx')
    wsqq=wb1q.active
    mx=wsqq.max_row
    Sheet1= wb1q[index_lstt[s2]]
    index_Data1=[]
    for i in range(3,mx+1,+1):
        index_Data=Sheet1['A'+str(i)].value
        index_Data1.append(index_Data)
    wb= Workbook()
    Vega= wb.active
    Vega['A1']='Scrip'
    Vega['B1']='Volatility'

    data_Vol= len(index_Data1)
    vol= data_Vol+1
    url='https://www.nseindia.com'
    scrip_List=[]
    scrip_List=index_Data1.copy()

    date_1= input('START DATE: dd-mm-yyyy: ')
    date_2= input('END DATE: dd-mm-yyyy: ')    

    print('\n')
    print('[IMPLIED VOLATILITY TABLE]'+'['+date_1+']'+'['+date_2+']')
    print('--------------------------------------------------------')
    print('\n')
    
    


    for SCRIP in scrip_List:
        url_Main = "https://www.nseindia.com/api/historical/cm/equity?symbol=" + SCRIP + "&series=[%22EQ%22]&from="+date_1+"&to="+date_2+"&csv=true"
        headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36 Edg/86.0.622.69'}
        session = requests.Session()
        request = session.get(url, headers=headers)
        cookies = dict(request.cookies)
        response = session.get(url_Main, headers=headers, cookies=cookies)
        req_Content= response.content
        try:
            csv_file=open( SCRIP +'.csv','wb')
            csv_file.write(req_Content)
            csv_file.close()
            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            with open(SCRIP +'.csv') as f:
                reader = csv.reader(f, delimiter=',')
                
                for row in reader:
                    ws2.append(row)
                
                reader = csv.reader(f, delimiter=':')
                for coloums in reader:
                    ws2.append(coloums)
            
            wb2.save(SCRIP+'.xlsx')
            wb1= openpyxl.load_workbook( SCRIP +'.xlsx')
            ws=wb1.active
            mx=ws.max_row
            sheet1= wb1['Sheet']
            data1=[]
            for j in range(mx,1,-1):
                l= sheet1['G' +str(j)].value
                m=l.replace(',','')
                data1.append(float(m))
                

            data2=[]    
            for k in range(0,mx-2):
                a=data1[k+1]
                b=data1[k]
                c=(a/b)-1
                data2.append(c)
            
            try:
                z= (statistics.stdev(data2))
                z1= round((z*100),2)
                z2= round((z*100*math.sqrt(30)),2)
                z3= round((z*100*math.sqrt(365)),2)
                z4= round((z*100*1.5*math.sqrt(30)),2)
                ltp1=data1[0]
                ltp2=data1[mx-2]
                return1= ((data1[mx-2])/(data1[0]))-1
                return2= return1*100
                return3=round(return2,2)
                return4= str(return3)+'%'
                return5=round(return3/z2,2)
                
                print(SCRIP)
                print('['+SCRIP+']'+'['+str(ltp1)+']'+'['+str(ltp2)+']')
                print('DAILY: '+str(z1)+'%')
                print('MONTHLY: '+str(z2)+'%')
                print('MONTHLY-1.5: '+str(z4)+'%')
                print('ANNUAL: '+str(z3)+'%')
                print('MONTHLY RAR: '+str(return5))
                print('['+SCRIP+']'+' has given a return of '+'['+return4+'] in this period.\n')
                
                
                
                os.remove(SCRIP+'.csv')
                os.remove(SCRIP+'.xlsx')
            except:
                pass
        except:
            pass

    wb.save('temp.xlsx')
    os.remove('temp.xlsx')
    print('--------------------------------------------------------')
    print('\n\n') 